import openpyxl


class _Excel:
    def __init__(self, excel_file_path, excel_sheet):
        self._m_wb = openpyxl.load_workbook(excel_file_path)
        self._m_sheet = self._m_wb[excel_sheet]

        self.__m_info = {}
        self.__m_info["max_row"] = self._m_sheet.max_row
        self.__m_info["max_col"] = self._m_sheet.max_column

    @property
    def _info(self):
        return self.__m_info


from enum import Enum


class ExcelTableType(Enum):
    ROW_BASED = 1
    COL_BASED = 2


from .tag import DataLabels


class ExcelTableConverter(_Excel):
    def __init__(self, type: ExcelTableType, excel_file_path, excel_sheet):
        super().__init__(excel_file_path, excel_sheet)
        self.__m_type = type

    def __iter__(self):
        self.__m_forsearch_index = 1
        if self.__m_forsearch_index == self.__m_datum_position:
            self.__m_forsearch_index += 1
        return self

    def __next__(self):
        if (self.__m_type == ExcelTableType.ROW_BASED and self.__m_forsearch_index >= self._info["max_row"]
        ) or (self.__m_type == ExcelTableType.COL_BASED and self.__m_forsearch_index >= self._info["max_col"]):
            raise StopIteration

        value = self.get_data(self.__m_forsearch_index)
        self.__m_forsearch_index += 1
        if self.__m_forsearch_index == self.__m_datum_position:
            self.__m_forsearch_index += 1
        return value

    def __get_one(self, position):
        get_one = []
        if self.__m_type == ExcelTableType.ROW_BASED:
            for columns in self._m_sheet.iter_cols(min_row=position, max_row=position, values_only=True):
                get_one.append(columns[0])
        elif self.__m_type == ExcelTableType.COL_BASED:
            for rows in self._m_sheet.iter_cols(min_col=position, max_col=position, values_only=True):
                get_one.append(rows[0])
        return get_one

    def set_datum(self, position, data_labels: DataLabels):
        self.__m_datum_position = position
        self.__m_data_labels = data_labels

        datum_position = self.__get_one(position)

        for label in data_labels:
            label.serial_num = datum_position.index(label.original)

    def get_data(self, position):
        python_dict = {}

        get_data = self.__get_one(position)

        for data_labels_index, associated_column_serial_num in zip(range(0, len(self.__m_data_labels)),
                                                                   self.__m_data_labels.associated_sequence):
            python_dict[str(self.__m_data_labels[data_labels_index])] = get_data[associated_column_serial_num]

        return python_dict

    @property
    def data_count(self):
        if self.__m_type == ExcelTableType.ROW_BASED:
            return self._info["max_row"] - 1
        elif self.__m_type == ExcelTableType.COL_BASED:
            return  self._info["max_col"] - 1
