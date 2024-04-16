import os.path
from typing import Tuple, List

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from data_packer.DataTag import DataTag
from data_packer.DataPack import DataPack


class Excel:
    def __init__(self, file_path: str, sheet_name: str):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file <{file_path}> does not exist.")
        else:
            self.__file_path: str = file_path

            # Get the Worksheet object using openpyxl.
            excel: Workbook = openpyxl.load_workbook(file_path)
            self.__excel: Worksheet = excel[sheet_name]

            # Getting the Excel sheet header.
            self.__header_tags: Tuple[str, ...] = tuple(str(cell.value) for cell in self.__excel[1])

            # Getting the number of rows in the Excel sheet.
            self.__long: int = self.__excel.max_row

    def __str__(self):
        return self.__file_path

    def get_row_data(self, row_index: int, matched_tags: List[DataTag]) -> Tuple[DataPack, ...]:
        matched_header_tag_indices: List[int] = []
        used_tag_names: List[str] = []

        for matched_tag in matched_tags:
            matched_header_tag_names: List[str] = []

            for matched_tag_name in matched_tag.names:
                if matched_tag_name in self.__header_tags:
                    matched_header_tag_names.append(matched_tag_name)

            if len(matched_header_tag_names) == 0:
                matched_header_tag_indices.append(-1)
            elif len(matched_header_tag_names) == 1:
                matched_header_tag_indices.append(self.__header_tags.index(matched_header_tag_names[0]) + 1)
                used_tag_names.append(matched_header_tag_names[0])
            else:
                raise ExcelHeaderTagConflictError(self, matched_header_tag_names, matched_tag)

        row_data: List[DataPack] = []
        matched_tags_index: int = 0

        for matched_header_tag_index in matched_header_tag_indices:
            matched_tag: DataTag = matched_tags[matched_tags_index]

            if matched_header_tag_index != -1:
                cell: Cell = self.__excel.cell(row=row_index, column=matched_header_tag_index)
                row_data.append(DataPack(matched_tag, used_tag_names[matched_tags_index], cell.value))
            else:
                row_data.append(DataPack(matched_tag, None, None))

            matched_tags_index += 1

        return tuple(row_data)

    @property
    def long(self) -> int:
        return self.__long


class ExcelHeaderTagConflictError(Exception):
    def __init__(self, excel: Excel, conflicting_header_tags: List[str], shared_data_tag: DataTag):
        self.__excel: Excel = excel
        self.__conflicting_header_tags: List[str] = conflicting_header_tags
        self.__shared_data_tag: DataTag = shared_data_tag
        super().__init__(f"The header tags <{','.join(self.__conflicting_header_tags)}> in excel <{self.__excel}> "
                         f"share the same DataTag {repr(self.__shared_data_tag)}.")
