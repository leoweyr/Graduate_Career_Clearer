import os.path
from typing import Tuple, Dict, Any, List

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell

from data_packer.DataType import DataType


class Excel:
    def __init__(self, file_path: str, sheet_name: str):
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"The file <{file_path}> does not exist.")
        else:
            self.__file_path: str = file_path

            # Get the Worksheet object using openpyxl.
            excel: Workbook = openpyxl.load_workbook(file_path)
            self.__excel: Worksheet = excel[sheet_name]

            # Getting the Excel sheet header.
            self.__header_tags: Tuple[str, ...] = tuple(str(cell.value) for cell in self.__excel[1])

            # Getting the number of rows in the Excel sheet.
            self.__long: int = self.__excel.max_row

    def get_row_data(self, row_index: int, matched_header_tags: Dict[str, DataType]) -> Tuple[Any, ...]:
        matched_header_tag_names: List[str] = list(matched_header_tags.keys())
        matched_header_tag_types: List[DataType] = list(matched_header_tags.values())
        matched_header_tag_indices: List[int] = []

        for matched_header_tag_name in matched_header_tag_names:
            if matched_header_tag_name in self.__header_tags:
                matched_header_tag_indices.append(self.__header_tags.index(matched_header_tag_name) + 1)

        row_data: List[Any] = []
        index: int = 0

        for matched_header_tag_index in matched_header_tag_indices:
            cell: Cell = self.__excel.cell(row=row_index, column=matched_header_tag_index)

            if matched_header_tag_types[index] == DataType.INTEGER:
                row_data.append(int(cell.value))
            elif matched_header_tag_types[index] == DataType.FLOAT:
                row_data.append(float(cell.value))
            elif matched_header_tag_types[index] == DataType.STRING:
                row_data.append(str(cell.value))

            index += 1

        return tuple(row_data)

    @property
    def long(self) -> int:
        return self.__long
